#!/usr/bin/env python3
"""
从500.com抓取北单数据，生成Excel看板（单表，合并让球胜平负+胜负过关）
支持自动获取当前期数或 --expect 参数指定

2026-09-03 数据源迁移: zx.500.com 被腾讯云 EdgeOne 防护封锁(bot挑战+人机验证码),
改为 trade.500.com (bjdc=让球胜平负 / bjdcsf=胜负过关) —— 该域未被防护.
注意: trade 源不含 zx 特有的"皇冠历史相同亚盘"概率列(胜%/平%/负%), 该列暂空.
"""
import re, os, sys, argparse
from datetime import datetime
from urllib.request import Request, urlopen
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJ_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# 2026-09-01 方案B: 北单产物(output)挪出主仓库docs/, 放独立目录 ~/beidan/, 避免混入football-dashboard
OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "beidan")
os.makedirs(OUTPUT_DIR, exist_ok=True)
TODAY = datetime.now().strftime("%Y-%m-%d")

# ===== Styles =====
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
HOME_FAV_FILL = PatternFill("solid", fgColor="E2EFDA")
AWAY_FAV_FILL = PatternFill("solid", fgColor="FCE4D6")

# 500.com 赛果列: 3=胜 1=平 0=负 *=无效场次
RESULT_MAP = {"3": "胜", "1": "平", "0": "负", "*": "＊"}


def fetch(url):
    req = Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://trade.500.com/bjdc/",
    })
    with urlopen(req, timeout=15) as resp:
        raw = resp.read()
    try:
        return raw.decode("gb2312")
    except:
        return raw.decode("gbk", errors="replace")


def strip_rank_home(s):
    """'[6] 奥卢' -> '奥卢'; 无排名则原样去前缀"""
    m = re.match(r'^\[(\d+)\]\s*(.+)$', s)
    return m.group(2).strip() if m else re.sub(r'^\s*\[\d+\]\s*', '', s).strip()


def strip_rank_away(s):
    """'塞那乔其 [10]' -> '塞那乔其'; 无排名则原样去后缀"""
    return re.sub(r'\s*\[\d+\]\s*$', '', s).strip()


def detect_current_expect():
    """Auto-detect current 期数 from trade.500.com dropdown (取最大值=当前期)"""
    print("🔍 自动检测当前期数...")
    text = fetch("https://trade.500.com/bjdc/")
    opts = [int(x) for x in re.findall(r'<option[^>]*value="?(\d{5})"?', text) if x.isdigit()]
    if opts:
        expect = max(opts)
        print(f"   ✓ 当前期数: {expect}")
        return expect
    raise RuntimeError("无法检测当前北单期数")


def fetch_page(playid, expect):
    if playid == 3:
        url = f"https://trade.500.com/bjdc/?expect={expect}"
    else:
        url = f"https://trade.500.com/bjdcsf/?expect={expect}"
    return fetch(url)


def parse_matches(text, playid):
    """Parse table rows. playid=3 -> 让球胜平负(bjdc); 默认 playid=0 -> 胜负过关(bjdcsf)"""
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', text, re.DOTALL)
    matches = {}
    for r in rows:
        cells = re.findall(r'<td[^>]*>(.*?)</td>', r, re.DOTALL)
        if len(cells) < 6:
            continue
        vals = []
        for c in cells:
            clean = re.sub(r'<[^>]+>', ' ', c)
            clean = re.sub(r'&nbsp;|\s+', ' ', clean).strip()
            vals.append(clean)
        first = vals[0].strip()
        if not first.isdigit():
            continue
        num = int(first)

        if playid == 3:
            # 让球胜平负(bjdc): [0]num [1]league [2]time [3]主队[排名] [4]让球 [5]客队[排名] [6]欧赔 [7]比分
            if len(vals) >= 8:
                score = vals[7] if re.match(r'^\d+:\d+$', vals[7]) else ""
                matches[num] = {
                    "num": num, "league": vals[1], "time": vals[2],
                    "home": strip_rank_home(vals[3]), "away": strip_rank_away(vals[5]),
                    "handicap": vals[4], "score": score, "result": "",
                    "h_prob": -1, "d_prob": -1, "a_prob": -1, "ah_desc": "",
                }
        else:
            # 胜负过关(bjdcsf): [0]num [1]足球 [2]league [3]time [4]主队 盘口 水位 客队 [5]比分 [6]欧赔
            if len(vals) >= 7:
                m = re.match(
                    r'^\s*\[?(\d*)\]?\s*(.+?)\s+([+-]?\d+\.?\d*球)\s+([\d.]+)(?:\s+([\d.]+))?\s+(.+?)(?:\s*\[\d+\])?\s*$',
                    vals[4].strip())
                if m:
                    time_tokens = vals[3].split()
                    score = vals[5] if re.match(r'^\d+:\d+$', vals[5]) else ""
                    matches[num] = {
                        "num": num, "league": vals[2],
                        "time": time_tokens[1] if len(time_tokens) > 1 else vals[3],
                        "home": m.group(2).strip(), "away": m.group(6).strip(),
                        "handicap": "", "score": score, "result": "",
                        "h_prob": -1, "d_prob": -1, "a_prob": -1,
                        "ah_desc": m.group(3), "water": m.group(4),
                    }
    return matches


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if r % 2 == 0:
                cell.fill = LIGHT_GRAY


def auto_width(ws, max_col, max_row, min_w=8, max_w=30):
    for col in range(1, max_col + 1):
        longest = min_w
        for row in range(1, max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                s = str(val)
                w = sum(2 if ord(c) > 127 else 1 for c in s)
                longest = max(longest, min(w + 2, max_w))
        ws.column_dimensions[get_column_letter(col)].width = longest


def prob_cell(ws, row, col, prob_val):
    cell = ws.cell(row=row, column=col)
    if prob_val >= 0:
        cell.value = f"{prob_val}%"
        if prob_val >= 45:
            cell.fill = GREEN_FILL
        elif prob_val >= 35:
            cell.fill = YELLOW_FILL


def main():
    parser = argparse.ArgumentParser(description="北单数据看板生成")
    parser.add_argument("--expect", type=int, default=None, help="北单期数，默认自动检测")
    args = parser.parse_args()

    expect = args.expect if args.expect else detect_current_expect()
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"beidan_{expect}_dashboard.xlsx")

    print(f"🔄 抓取北单{expect}期数据...")

    m3 = parse_matches(fetch_page(3, expect), 3)
    m0 = parse_matches(fetch_page(0, expect), 0)
    print(f"   ✓ 让球胜平负: {len(m3)} 场  胜负过关: {len(m0)} 场")

    all_nums = sorted(set(list(m3.keys()) + list(m0.keys())))
    print(f"   ✓ 合并后共 {len(all_nums)} 场")

    wb = Workbook()
    ws = wb.active
    ws.title = f"北单{expect}期"

    headers = [
        "序号", "联赛", "开赛时间", "主队", "让球(胜平负)", "客队",
        "比分", "赛果",
        "亚盘(胜负过关)", "皇冠历史概率% (主胜)", "皇冠历史概率% (平)", "皇冠历史概率% (客胜)"
    ]
    max_col = len(headers)

    ws.cell(row=1, column=1, value=f"⚽ 北京单场{expect}期 比赛看板 — {TODAY}").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    ws.cell(row=2, column=1, value="数据来源：500.com trade（让球胜平负+胜负过关）│ 皇冠历史概率列暂缺（zx源被EdgeOne封锁，需另源）").font = Font(
        name="微软雅黑", size=9, italic=True, color="888888"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, max_col)

    row = 4
    for num in all_nums:
        d3 = m3.get(num, {})
        d0 = m0.get(num, {})

        league = d3.get("league", d0.get("league", ""))
        time_val = d3.get("time", d0.get("time", ""))
        home = d3.get("home", d0.get("home", ""))
        away = d3.get("away", d0.get("away", ""))
        hdcp = d3.get("handicap", d0.get("handicap", ""))

        # Use playid=0 probs (胜负过关) if available, else playid=3
        if d0.get("h_prob", -1) >= 0:
            h_prob, d_prob, a_prob = d0["h_prob"], d0["d_prob"], d0["a_prob"]
        else:
            h_prob, d_prob, a_prob = d3.get("h_prob", -1), d3.get("d_prob", -1), d3.get("a_prob", -1)

        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=league)
        ws.cell(row=row, column=3, value=time_val)
        ws.cell(row=row, column=4, value=home)

        hdcp_cell = ws.cell(row=row, column=5, value=hdcp)
        try:
            hnum = int(hdcp)
            if hnum < 0:
                hdcp_cell.fill = HOME_FAV_FILL
            elif hnum > 0:
                hdcp_cell.fill = AWAY_FAV_FILL
        except:
            pass

        ws.cell(row=row, column=6, value=away)
        score = d3.get("score") or d0.get("score") or ""
        result = RESULT_MAP.get(d3.get("result") or d0.get("result") or "", "")
        ws.cell(row=row, column=7, value=score)
        ws.cell(row=row, column=8, value=result)
        ws.cell(row=row, column=9, value=d0.get("ah_desc", ""))
        prob_cell(ws, row, 10, h_prob)
        prob_cell(ws, row, 11, d_prob)
        prob_cell(ws, row, 12, a_prob)

        row += 1

    end_row = row - 1
    style_body(ws, 4, end_row, max_col)
    auto_width(ws, max_col, end_row)

    row += 1
    ws.cell(row=row, column=1, value=f"共 {end_row - 3} 场比赛").font = Font(
        name="微软雅黑", size=10, italic=True, color="666666"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    row += 2
    ws.cell(row=row, column=1, value="图例：").font = BOLD_FONT
    row += 1
    for fill, desc in [
        (HOME_FAV_FILL, "让球为负（主队让球）"),
        (AWAY_FAV_FILL, "让球为正（客队让球）"),
        (GREEN_FILL, "历史概率≥45%"),
        (YELLOW_FILL, "历史概率≥35%"),
    ]:
        c = ws.cell(row=row, column=1, value="  ")
        c.fill = fill
        c.border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).font = BODY_FONT
        row += 1

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 看板已生成: {OUTPUT_FILE}")
    print(f"\n📋 预览 (前5场):")
    for num in all_nums[:5]:
        d3, d0 = m3.get(num, {}), m0.get(num, {})
        l = d3.get("league", d0.get("league", ""))
        t = d3.get("time", d0.get("time", ""))
        h = d3.get("home", d0.get("home", ""))
        a = d3.get("away", d0.get("away", ""))
        hc = d3.get("handicap", d0.get("handicap", ""))
        ah = d0.get("ah_desc", "")
        hp, dp, ap = d0.get("h_prob", -1), d0.get("d_prob", -1), d0.get("a_prob", -1)
        sc = d3.get("score") or d0.get("score") or "-"
        rs = RESULT_MAP.get(d3.get("result") or d0.get("result") or "", "")
        probstr = (f"胜{hp}%平{dp}%负{ap}%" if (hp >= 0 and dp >= 0 and ap >= 0) else "皇冠概率待补")
        print(f"   {num}. {l} {t} {h}({hc}){a} [{ah}] {probstr} 比分{sc} {rs}")

    return OUTPUT_FILE


if __name__ == "__main__":
    main()
